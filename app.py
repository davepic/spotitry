import os
from flask import Flask, render_template, request, redirect, session, url_for, send_from_directory
from werkzeug import secure_filename
from flask.ext.mongoengine import MongoEngine
from flask.ext.login import LoginManager, login_user, logout_user, login_required, current_user
from flask.ext.mongoengine.wtf import model_form
from wtforms import PasswordField
import datetime
import urllib
import base64
import json
from datetime import *
from flask_mail import Mail, Message
import requests.packages.urllib3
from flask.ext.paginate import Pagination
from openpyxl import *

requests.packages.urllib3.disable_warnings()

UPLOAD_FOLDER = '/Users/Dave/Documents'
ALLOWED_EXTENSIONS = set(['txt', 'xlsx'])



app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config["DEBUG"] = True
app.config['SECRET_KEY'] = 'ajosjdfajjjj3453453oj!!!oij'
app.config['WTF_CSRF_ENABLED'] = True


CLIENT_ID = "68eb0b1766a64f8294da279e662232c0"
CLIENT_SECRET = "8a2d182155d64971846f16faae25f1e4"

SPOTIFY_AUTH_URL = "https://accounts.spotify.com/authorize"
SPOTIFY_TOKEN_URL = "https://accounts.spotify.com/api/token"
SPOTIFY_API_BASE_URL = "https://api.spotify.com"
API_VERSION = "v1"
SPOTIFY_API_URL = "{}/{}".format(SPOTIFY_API_BASE_URL, API_VERSION)



CLIENT_SIDE_URL = "http://0.0.0.0"
PORT = 5000
SCOPE = "playlist-modify-public playlist-modify-private user-library-modify"
STATE = ""
SHOW_DIALOG_bool = True
SHOW_DIALOG_str = str(SHOW_DIALOG_bool).lower()



def allowed_file(filename):
	return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


@app.route("/playlist", methods=["POST", "GET"])
def playlist():

	REDIRECT_URI = "{}:{}/playlists/q".format(CLIENT_SIDE_URL, PORT)

	auth_query_parameters = {
    "response_type": "code",
    "redirect_uri": REDIRECT_URI,
    "scope": SCOPE,
    "client_id": CLIENT_ID
	}

	url_args = "&".join(["{}={}".format(key,urllib.quote(val)) for key,val in auth_query_parameters.iteritems()])
	auth_url = "{}/?{}".format(SPOTIFY_AUTH_URL, url_args)
	return redirect(auth_url)

@app.route("/setlist", methods=["POST", "GET"])
def setlist():


	if request.method == "POST":

		REDIRECT_URI = "{}:{}/setlists/q".format(CLIENT_SIDE_URL, PORT)

		session["setlist_name"] = request.form["playlist"]
		session["setlist_id"] = request.form["setlist_id"]

		auth_query_parameters = {
    	"response_type": "code",
    	"redirect_uri": REDIRECT_URI,
    	"scope": SCOPE,
    	"client_id": CLIENT_ID
		}

		url_args = "&".join(["{}={}".format(key,urllib.quote(val)) for key,val in auth_query_parameters.iteritems()])
		auth_url = "{}/?{}".format(SPOTIFY_AUTH_URL, url_args)
		
		return redirect(auth_url)

	return render_template("setlister.html")



@app.route("/upload", methods=["POST", "GET"])
def upload():

	REDIRECT_URI = "{}:{}/upload/q".format(CLIENT_SIDE_URL, PORT)

	auth_query_parameters = {
    "response_type": "code",
    "redirect_uri": REDIRECT_URI,
    "scope": SCOPE,
    "client_id": CLIENT_ID
	}

	url_args = "&".join(["{}={}".format(key,urllib.quote(val)) for key,val in auth_query_parameters.iteritems()])
	auth_url = "{}/?{}".format(SPOTIFY_AUTH_URL, url_args)
	return redirect(auth_url)

@app.route("/setlists/q", methods=["POST", "GET"])
def setlisted():

	setlist = []
	songs= []
	errors = []

	REDIRECT_URI = "{}:{}/setlists/q".format(CLIENT_SIDE_URL, PORT)

	auth_token = request.args['code']
	code_payload = {
    	"grant_type": "authorization_code",
        "code": str(auth_token),
        "redirect_uri": REDIRECT_URI
    }
	base64encoded = base64.b64encode("{}:{}".format(CLIENT_ID, CLIENT_SECRET))
	headers = {"Authorization": "Basic " + base64encoded}
	post_request = requests.post(SPOTIFY_TOKEN_URL, data=code_payload, headers=headers)

	# Auth Step 5: Tokens are Returned to Application
	response_data = json.loads(post_request.text)
	access_token = response_data["access_token"]
	refresh_token = response_data["refresh_token"]
	token_type = response_data["token_type"]
	expires_in = response_data["expires_in"]

    # Auth Step 6: Use the access token to access Spotify API
	authorization_header = {"Authorization":"Bearer " + access_token, "Content-Type": "application/json"}

	# Get profile data
	user_profile_api_endpoint = "{}/me".format(SPOTIFY_API_URL)
	profile_data = requests.get(user_profile_api_endpoint, headers=authorization_header).json()

	playlist_api_endpoint = "{}/playlists".format(profile_data["href"])


	playlist_create_response = requests.post(playlist_api_endpoint, data="{\"name\":\"" + session.pop("setlist_name") + "\"}", headers=authorization_header).json()
	
	playlist_url = playlist_create_response["tracks"]["href"]
	

	setlist_url = "http://api.setlist.fm/rest/0.1/setlist/" + session.pop("setlist_id") + ".json"



	index = 0


	try:
		setlist_dict = requests.get(setlist_url).json()

		
		if setlist_dict["setlist"]["sets"] != "":

			artist_name = setlist_dict["setlist"]["artist"]["@name"]

			if type(setlist_dict["setlist"]["sets"]["set"]	) is list:
				for show_setlist in setlist_dict["setlist"]["sets"]["set"]:
					if "@encore" in show_setlist.keys():
						setlist.append(("Encore " + show_setlist["@encore"]+" :", ""))
						
					if type(show_setlist["song"]) is list:
						for song in show_setlist["song"]:
							if "cover" in song.keys():
								setlist.append((song["@name"], song["cover"]["@name"]))
							else:
								setlist.append((song["@name"], ""))
							
					else:
						if "cover" in show_setlist["song"].keys():
							setlist.append((show_setlist["song"]["@name"], show_setlist["song"]["cover"]["@name"])) 
						else:
							setlist.append((show_setlist["song"]["@name"], ""))
						

			else:
				for song in setlist_dict["setlist"]["sets"]["set"]["song"]:
					if "cover" in song.keys():
						setlist.append((song["@name"], song["cover"]["@name"]))
					else:
						setlist.append((song["@name"], ""))



		for song in setlist:
			if song[0][:6] != "Encore":

				try:
					track_info = requests.get("https://api.spotify.com/v1/search?q=track:" + song[0] +"%20artist:" + artist_name + "&type=track").json()
					
					if track_info["tracks"]["items"] == []:
						
						if song[1] != "":
							#track_info = requests.get("https://api.spotify.com/v1/search?q=track:" + song + "&type=track").json()
							track_info = requests.get("https://api.spotify.com/v1/search?q=track:" + song[0] +"%20artist:" + song[1] + "&type=track").json()
						

							if track_info["tracks"]["items"] == []:
								errors.append(song)
							else:
								song_response = requests.post(playlist_url + "?uris=" + track_info["tracks"]["items"][0]["uri"], headers=authorization_header).json()
								songs.append(song)
						else:
							errors.append(song)
					else:
						song_response = requests.post(playlist_url + "?uris=" + track_info["tracks"]["items"][0]["uri"], headers=authorization_header).json()
						songs.append(song)

				except ValueError:
					errors.append(song)
					
	except ValueError:
		pass

	#for i in range(len(session["song_list"])):
	#	song_endpoint = "https://api.spotify.com/v1/search?q=track:" + session["song_list"][i][0]+  "%20artist:" + session["song_list"][i][1]+ "&type=track"
	#	song_data = requests.get(song_endpoint).json()
	#	if song_data["tracks"]["items"] != []:
	#		song_response = requests.post(song_url + "?uris=" + song_data["tracks"]["items"][0]["uri"], headers=authorization_header).json()

	
	return render_template("setlister.html")

@app.route("/upload/q", methods=["POST", "GET"])
def uploaded():

	REDIRECT_URI = "{}:{}/upload/q".format(CLIENT_SIDE_URL, PORT)

	auth_token = request.args['code']
	code_payload = {
    	"grant_type": "authorization_code",
        "code": str(auth_token),
        "redirect_uri": REDIRECT_URI
    }
	base64encoded = base64.b64encode("{}:{}".format(CLIENT_ID, CLIENT_SECRET))
	headers = {"Authorization": "Basic " + base64encoded}
	post_request = requests.post(SPOTIFY_TOKEN_URL, data=code_payload, headers=headers)

	# Auth Step 5: Tokens are Returned to Application
	response_data = json.loads(post_request.text)
	access_token = response_data["access_token"]
	refresh_token = response_data["refresh_token"]
	token_type = response_data["token_type"]
	expires_in = response_data["expires_in"]

    # Auth Step 6: Use the access token to access Spotify API
	authorization_header = {"Authorization":"Bearer " + access_token, "Content-Type": "application/json"}

	# Get profile data
	user_profile_api_endpoint = "{}/me".format(SPOTIFY_API_URL)
	profile_data = requests.get(user_profile_api_endpoint, headers=authorization_header).json()

	playlist_api_endpoint = "{}/playlists".format(profile_data["href"])
	
	response = requests.post(playlist_api_endpoint, data="{\"name\":\"" + session.pop("playlist_name") + "\"}", headers=authorization_header).json()
	song_url = response["tracks"]["href"]

	for i in range(len(session["song_list"])):
		song_endpoint = "https://api.spotify.com/v1/search?q=track:" + session["song_list"][i][0]+  "%20artist:" + session["song_list"][i][1]+ "&type=track"
		song_data = requests.get(song_endpoint).json()
		if song_data["tracks"]["items"] != []:
			song_response = requests.post(song_url + "?uris=" + song_data["tracks"]["items"][0]["uri"], headers=authorization_header).json()

	
	return render_template("upload.html", song_list = session.pop("song_list"))



@app.route("/save", methods=["POST", "GET"])
def save():

	REDIRECT_URI = "{}:{}/saved/q".format(CLIENT_SIDE_URL, PORT)

	auth_query_parameters = {
    "response_type": "code",
    "redirect_uri": REDIRECT_URI,
    "scope": SCOPE,
    "client_id": CLIENT_ID
	}

	url_args = "&".join(["{}={}".format(key,urllib.quote(val)) for key,val in auth_query_parameters.iteritems()])
	auth_url = "{}/?{}".format(SPOTIFY_AUTH_URL, url_args)
	return redirect(auth_url)





@app.route("/saved/q", methods=["POST", "GET"])
def saved():

	REDIRECT_URI = "{}:{}/saved/q".format(CLIENT_SIDE_URL, PORT)


	auth_token = request.args['code']
	code_payload = {
    	"grant_type": "authorization_code",
        "code": str(auth_token),
        "redirect_uri": REDIRECT_URI
    }
	base64encoded = base64.b64encode("{}:{}".format(CLIENT_ID, CLIENT_SECRET))
	headers = {"Authorization": "Basic " + base64encoded}
	post_request = requests.post(SPOTIFY_TOKEN_URL, data=code_payload, headers=headers)

	# Auth Step 5: Tokens are Returned to Application
	response_data = json.loads(post_request.text)
	access_token = response_data["access_token"]
	refresh_token = response_data["refresh_token"]
	token_type = response_data["token_type"]
	expires_in = response_data["expires_in"]

    # Auth Step 6: Use the access token to access Spotify API
	authorization_header = {"Authorization":"Bearer " + access_token, "Content-Type": "application/json"}

	# Get profile data
	user_profile_api_endpoint = "{}/me".format(SPOTIFY_API_URL)
	profile_data = requests.get(user_profile_api_endpoint, headers=authorization_header).json()
	
    
	song_endpoint = "https://api.spotify.com/v1/search?q=track:" + session["current_song"][0]+  "%20artist:" + session["artist_name"][0]+ "&type=track"

	song_data = requests.get(song_endpoint).json()


	song_id = song_data["tracks"]["items"][0]["id"]

	request_data = "[\"" + song_id + "\"]"
	

	response = requests.put("https://api.spotify.com/v1/me/tracks", headers=authorization_header, data= request_data)


	return render_template("save.html",  saved_song= session.pop("current_song"), artist_name=session.pop("artist_name"), album_image= session.pop("album_image"))


@app.route("/delete", methods=["POST", "GET"])
def delete():

	session["playlist"] = request.form["playlist"]


	REDIRECT_URI = "{}:{}/callback/q".format(CLIENT_SIDE_URL, PORT)

	auth_query_parameters = {
    "response_type": "code",
    "redirect_uri": REDIRECT_URI,
    "scope": SCOPE,
    "client_id": CLIENT_ID
	}

	url_args = "&".join(["{}={}".format(key,urllib.quote(val)) for key,val in auth_query_parameters.iteritems()])
	auth_url = "{}/?{}".format(SPOTIFY_AUTH_URL, url_args)
	
	return redirect(auth_url)


@app.route("/playlists/q", methods=["POST", "GET"])
def playlists():

	REDIRECT_URI = "{}:{}/playlists/q".format(CLIENT_SIDE_URL, PORT)

	playlists = []
	playlist_dict = {}

	auth_token = request.args['code']
	code_payload = {
    	"grant_type": "authorization_code",
        "code": str(auth_token),
        "redirect_uri": REDIRECT_URI
    }
	base64encoded = base64.b64encode("{}:{}".format(CLIENT_ID, CLIENT_SECRET))
	headers = {"Authorization": "Basic " + base64encoded}
	post_request = requests.post(SPOTIFY_TOKEN_URL, data=code_payload, headers=headers)

	# Auth Step 5: Tokens are Returned to Application
	response_data = json.loads(post_request.text)
	access_token = response_data["access_token"]
	refresh_token = response_data["refresh_token"]
	token_type = response_data["token_type"]
	expires_in = response_data["expires_in"]

    # Auth Step 6: Use the access token to access Spotify API
	authorization_header = {"Authorization":"Bearer " + access_token, "Content-Type": "application/json"}

	# Get profile data
	user_profile_api_endpoint = "{}/me".format(SPOTIFY_API_URL)
	profile_data = requests.get(user_profile_api_endpoint, headers=authorization_header).json()
	
    #Get user playlist data
	playlist_api_endpoint = "{}/playlists".format(profile_data["href"])
	playlists_data = requests.get(playlist_api_endpoint, headers=authorization_header).json()

	for playlist in playlists_data["items"]:
		playlists.append((playlist["name"], playlist["id"]))
		playlist_dict[playlist["id"]] = playlist["name"]

	session["playlist_dict"] = playlist_dict

	return render_template("playlist.html", playlists = playlists, current_song= session["current_song"], artist_name=session["artist_name"], album_image= session["album_image"], now_playing = session["now_playing"])

@app.route("/callback/q", methods=["POST", "GET"])
def callback():


	playlist_id = {}
	playlist_name = ""

	REDIRECT_URI = "{}:{}/callback/q".format(CLIENT_SIDE_URL, PORT)
	auth_token = request.args['code']
	code_payload = {
    	"grant_type": "authorization_code",
        "code": str(auth_token),
        "redirect_uri": REDIRECT_URI
    }
	base64encoded = base64.b64encode("{}:{}".format(CLIENT_ID, CLIENT_SECRET))
	headers = {"Authorization": "Basic " + base64encoded}
	post_request = requests.post(SPOTIFY_TOKEN_URL, data=code_payload, headers=headers)

	# Auth Step 5: Tokens are Returned to Application
	response_data = json.loads(post_request.text)
	access_token = response_data["access_token"]
	refresh_token = response_data["refresh_token"]
	token_type = response_data["token_type"]
	expires_in = response_data["expires_in"]

    # Auth Step 6: Use the access token to access Spotify API
	authorization_header = {"Authorization":"Bearer " + access_token, "Content-Type": "application/json"}

    # Get profile data
	user_profile_api_endpoint = "{}/me".format(SPOTIFY_API_URL)
	profile_data = requests.get(user_profile_api_endpoint, headers=authorization_header).json()
	
    # Get user playlist data
	playlist_api_endpoint = "{}/playlists/{}/tracks".format(profile_data["href"], session["playlist"])


	song_endpoint = "https://api.spotify.com/v1/search?q=track:" + session["current_song"][0]+  "%20artist:" + session["artist_name"][0]+ "&type=track"


	song_data = requests.get(song_endpoint).json()

	song_uri = song_data["tracks"]["items"][0]["uri"]

	request_data = "{\"tracks\": [{\"uri\":\"" + song_uri+ "\"}]}"

	playlist_id = session.pop("playlist_dict")

	playlist_name = playlist_id[session.pop("playlist")]


	response = requests.delete(playlist_api_endpoint, data = request_data, headers= authorization_header).json()

	



	if response.get("snapshot_id"):
		return render_template("delete.html", success= True, playlist_name = playlist_name, deleted_song=session.pop("current_song"), album_image = session.pop("album_image"), artist_name = session.pop("artist_name"))
	else:
		return render_template("delete.html", success=False)




@app.route("/edit")
def edit():

	song_info = requests.get("http://ws.audioscrobbler.com/2.0/?method=user.getrecenttracks&user=dpiccolella&api_key=e8d959e5d2743dca3d2962338084b0ed&format=json").json()
	current_song = []
	album_image = []
	artist_name = []
	now_playing = False


	current_song.append(song_info["recenttracks"]["track"][0]["name"])
	album_image.append(song_info["recenttracks"]["track"][0]["image"][3]["#text"])

	artist_name.append(song_info["recenttracks"]["track"][0]["artist"]["#text"])
	

	if "@attr" in song_info["recenttracks"]["track"][0].keys():

		if song_info["recenttracks"]["track"][0].get("@attr")["nowplaying"] == "true":
			now_playing = True

	session["current_song"] = current_song
	session["album_image"] = album_image
	session["artist_name"] = artist_name
	session["now_playing"] = now_playing

	return redirect("/playlist")


@app.route("/uploads", methods = ['GET', 'POST'])
def uploads():


	song_list = []
	j=1


	if request.method == 'POST':
		file = request.files['file']
		if file and allowed_file(file.filename):
			filename = secure_filename(file.filename)

			wb = load_workbook(filename=filename)
			ws1 = wb.active



			while ws1.cell(column=1, row=j).value:

				song_list.append((str(ws1.cell(column=1, row=j).value).capitalize(), str(ws1.cell(column=2, row=j).value).capitalize()))

				j=j+1

			session["song_list"] = song_list
			session["playlist_name"] = request.form["playlist"]
		
			
			return redirect("/upload")
			


	return render_template("form.html")


@app.route("/")
def home():

	return render_template('base.html')

@app.route('/uploads/<filename>')
def uploaded_file(filename):
	return send_from_directory(app.config['UPLOAD_FOLDER'], filename)



if __name__ == "__main__":

	app.run(host="0.0.0.0")

